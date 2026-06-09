from __future__ import annotations

import os
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path.cwd()
ASSET_FILE = next(ROOT.glob("*.md"))
OUTPUT_DIR = ROOT / "outputs" / "may_2026_financial_reports"
OUTPUT_FILE = OUTPUT_DIR / "芒果租车2026年5月毛利及财务报表.xlsx"


def money(v):
    return round(float(v or 0), 2)


def find_source_workbook() -> str:
    text = ASSET_FILE.read_text(encoding="utf-8", errors="ignore")
    paths = list(dict.fromkeys(re.findall(r"C:\\\\[^`|<\\n\\r]+?\\.xls[x]?", text)))
    for raw_path in paths:
        p = raw_path.replace("\\\\", "\\")
        if (
            os.path.exists(p)
            and "2026" in os.path.basename(p)
            and "5" in os.path.basename(p)
            and os.path.getsize(p) == 85523
        ):
            return p
    search_roots = [Path.home() / "Desktop", Path.home() / "Documents", Path.home() / "Downloads"]
    for search_root in search_roots:
        if not search_root.exists():
            continue
        for candidate in search_root.rglob("*.xlsx"):
            try:
                if "2026" in candidate.name and "5" in candidate.name and candidate.stat().st_size == 85523:
                    return str(candidate)
            except OSError:
                continue
    raise FileNotFoundError("未找到 2026 年 5 月主收支表")


def includes_any(text: str, keywords: list[str]) -> bool:
    return any(k in text for k in keywords)


DIRECT_KEYWORDS = [
    "板车",
    "GPS",
    "定位",
    "违章",
    "保险",
    "加油",
    "油费",
    "收车",
    "送车",
    "配驾",
    "代驾",
    "停车",
    "过路",
    "高速",
    "物流",
    "运费",
    "洗车",
    "补油",
    "车主",
    "车辆",
]

OPEX_KEYWORDS = [
    "工资",
    "预支",
    "社保",
    "抖",
    "办公",
    "电费",
    "物业",
    "仲裁",
    "诉讼",
    "企业微信",
    "餐",
    "聚餐",
    "沙发",
    "扩容",
    "商标",
    "充值",
    "公户",
]

NON_OPERATING_RECEIPT_KEYWORDS = ["老板微信转入", "老板", "老板银行卡"]


def load_data(source_path: str):
    wb = load_workbook(source_path, data_only=True, read_only=True)
    receipt_ws = wb.worksheets[1]
    payment_ws = wb.worksheets[2]

    receipts = []
    for row in receipt_ws.iter_rows(min_row=3, values_only=True):
        values = list(row) + [None] * 7
        date, customer, method, typ, amount, summary, remark = values[:7]
        try:
            amount = float(amount)
        except (TypeError, ValueError):
            continue
        receipts.append(
            {
                "date": date,
                "customer": customer,
                "method": method,
                "type": typ,
                "amount": amount,
                "summary": summary,
                "remark": remark,
            }
        )

    payments = []
    for row in payment_ws.iter_rows(min_row=3, values_only=True):
        values = list(row) + [None] * 8
        date, payee, method, typ, amount, summary, remark, handler = values[:8]
        try:
            amount = float(amount)
        except (TypeError, ValueError):
            continue
        payments.append(
            {
                "date": date,
                "payee": payee,
                "method": method,
                "type": typ,
                "amount": amount,
                "summary": summary,
                "remark": remark,
                "handler": handler,
            }
        )
    return receipts, payments


def classify(receipts, payments):
    receipt_by_type = defaultdict(float)
    receipt_by_method = defaultdict(float)
    payment_by_type = defaultdict(float)
    payment_by_method = defaultdict(float)
    payment_by_handler = defaultdict(float)

    for r in receipts:
        receipt_by_type[r["type"] or "未分类"] += r["amount"]
        receipt_by_method[r["method"] or "未分类"] += r["amount"]
    for p in payments:
        payment_by_type[p["type"] or "未分类"] += p["amount"]
        payment_by_method[p["method"] or "未分类"] += p["amount"]
        payment_by_handler[p["handler"] or "未分类"] += p["amount"]

    boss_transfer = 0.0
    operating_other_receipts = 0.0
    other_receipt_rows = []
    for r in receipts:
        text = f'{r.get("customer") or ""} {r.get("summary") or ""} {r.get("remark") or ""}'
        if r["type"] == "其他":
            if includes_any(text, NON_OPERATING_RECEIPT_KEYWORDS):
                boss_transfer += r["amount"]
            else:
                operating_other_receipts += r["amount"]
                other_receipt_rows.append(r)

    revenue = receipt_by_type["租金"] + operating_other_receipts

    direct_payment_types = ["同行结算", "挂靠结算", "维修结算"]
    direct_cost = sum(payment_by_type[t] for t in direct_payment_types)
    opex = payment_by_type["工资、社保"]
    refund_outflow = payment_by_type["退客户押金"] + payment_by_type["退客户定金"]
    direct_extra = 0.0
    opex_extra = 0.0
    unclassified_extra = 0.0

    for p in payments:
        if p["type"] not in ["其他", "日常经营费用报销"]:
            continue
        text = f'{p.get("payee") or ""} {p.get("summary") or ""} {p.get("remark") or ""}'
        if includes_any(text, OPEX_KEYWORDS):
            opex_extra += p["amount"]
        elif includes_any(text, DIRECT_KEYWORDS):
            direct_extra += p["amount"]
        else:
            unclassified_extra += p["amount"]

    direct_cost += direct_extra
    opex += opex_extra + unclassified_extra
    gross_profit = revenue - direct_cost
    operating_profit = gross_profit - opex
    gross_margin = gross_profit / revenue if revenue else 0
    operating_margin = operating_profit / revenue if revenue else 0

    return {
        "receipt_by_type": dict(receipt_by_type),
        "receipt_by_method": dict(receipt_by_method),
        "payment_by_type": dict(payment_by_type),
        "payment_by_method": dict(payment_by_method),
        "payment_by_handler": dict(payment_by_handler),
        "boss_transfer": boss_transfer,
        "operating_other_receipts": operating_other_receipts,
        "other_receipt_rows": other_receipt_rows,
        "revenue": revenue,
        "direct_cost": direct_cost,
        "direct_extra": direct_extra,
        "opex": opex,
        "opex_extra": opex_extra,
        "unclassified_extra": unclassified_extra,
        "refund_outflow": refund_outflow,
        "gross_profit": gross_profit,
        "operating_profit": operating_profit,
        "gross_margin": gross_margin,
        "operating_margin": operating_margin,
        "cash_receipts": sum(receipt_by_type.values()),
        "receipt_type_total_excl_violation_deposit": sum(receipt_by_type.values()) - receipt_by_type["违章押金"],
        "cash_payments": sum(payment_by_type.values()),
        "net_cash_flow": sum(receipt_by_type.values()) - sum(payment_by_type.values()),
        "deposit_receipts": receipt_by_type["押金"] + receipt_by_type["违章押金"] + receipt_by_type["定金"],
    }


def style_sheet(ws, freeze="A2"):
    ws.freeze_panes = freeze
    ws.sheet_view.showGridLines = False
    thin = Side(style="thin", color="D9E2EC")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = Border(bottom=thin)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="17324D")
        cell.alignment = Alignment(horizontal="center", vertical="center")


def set_widths(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def write_table(ws, start_row, headers, rows, number_cols=None):
    number_cols = set(number_cols or [])
    for c, h in enumerate(headers, 1):
        cell = ws.cell(start_row, c, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="315A7D")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r_idx, row in enumerate(rows, start_row + 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(r_idx, c_idx, value)
            if c_idx in number_cols:
                cell.number_format = '#,##0.00;[Red]-#,##0.00;-'
                cell.alignment = Alignment(horizontal="right")
    return start_row + len(rows)


def build_workbook(source_path, receipts, payments, metrics):
    wb = Workbook()
    wb.remove(wb.active)

    title_fill = PatternFill("solid", fgColor="17324D")
    blue_fill = PatternFill("solid", fgColor="E8F1F8")
    green_fill = PatternFill("solid", fgColor="E6F4EA")
    warn_fill = PatternFill("solid", fgColor="FFF4CE")
    bad_fill = PatternFill("solid", fgColor="FCE8E6")

    # Cover
    ws = wb.create_sheet("封面与口径")
    rows = [
        ["报表名称", "芒果租车2026年5月毛利及财务报表"],
        ["期间", "2026年5月1日 - 2026年5月31日"],
        ["主数据源", source_path],
        ["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["核心口径", "营业收入=租金+经营性其他收款；押金/违章押金/定金、老板转入不计入收入。"],
        ["毛利口径", "毛利=营业收入-同行结算-挂靠结算-维修结算-订单直接费用。"],
        ["注意事项", "原始表存在收款类型与摘要不完全一致的情况，本报表保留了口径说明和核对页，便于复核。"],
    ]
    write_table(ws, 1, ["项目", "说明"], rows, [])
    set_widths(ws, {"A": 18, "B": 120})
    style_sheet(ws)

    # Boss report
    ws = wb.create_sheet("老板版报表")
    summary_rows = [
        ["营业收入", metrics["revenue"], "租金 + 经营性其他收款，不含押金/定金/老板转入"],
        ["直接成本", metrics["direct_cost"], "同行、挂靠、维修及订单直接费用"],
        ["毛利", metrics["gross_profit"], "营业收入 - 直接成本"],
        ["毛利率", metrics["gross_margin"], "毛利 / 营业收入"],
        ["期间费用", metrics["opex"], "工资社保、营销、办公、仲裁、电费等"],
        ["经营利润（现金口径）", metrics["operating_profit"], "毛利 - 期间费用"],
        ["经营利润率", metrics["operating_margin"], "经营利润 / 营业收入"],
        ["收款总额", metrics["cash_receipts"], "现金流入总额，含押金/定金/老板转入"],
        ["支出总额", metrics["cash_payments"], "现金流出总额，含退押/退定"],
        ["净现金流", metrics["net_cash_flow"], "收款总额 - 支出总额"],
    ]
    ws.append(["老板版经营看板", "金额/比例", "说明"])
    for row in summary_rows:
        ws.append(row)
    for r in range(2, 12):
        ws.cell(r, 2).number_format = '0.0%' if "率" in str(ws.cell(r, 1).value) else '#,##0.00;[Red]-#,##0.00;-'
    ws["A14"] = "管理建议"
    ws["A14"].font = Font(bold=True, size=13)
    advice = [
        "1. 收款总额较高，但押金、定金和老板转入占用较多，建议老板看经营利润时不要直接用现金流替代利润。",
        "2. 毛利率是本月最核心指标；后续应把租金中“租金+押金”混写的摘要逐单拆清，提升毛利准确度。",
        "3. 直接费用中板车、收送车、加油/配驾等建议按订单或车牌归集，方便判断高收入车型是否真赚钱。",
        "4. 老板转入本月已剔除营业收入，作为资金补充列示，避免对外展示时被误认为业务收入。",
    ]
    for i, text in enumerate(advice, 15):
        ws.cell(i, 1, text)
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=3)
    set_widths(ws, {"A": 26, "B": 18, "C": 80})
    style_sheet(ws)
    for row in [2, 4, 7, 10]:
        ws.cell(row, 1).fill = green_fill
        ws.cell(row, 2).fill = green_fill

    # Shareholder report
    ws = wb.create_sheet("股东版报表")
    rows = [
        ["营业收入", metrics["revenue"], "以已入账经营性收款为基础，剔除押金、定金及老板资金转入"],
        ["毛利", metrics["gross_profit"], "扣除直接车辆/订单成本后利润"],
        ["毛利率", metrics["gross_margin"], "体现租车业务单月盈利质量"],
        ["经营利润（现金口径）", metrics["operating_profit"], "扣除期间费用后的经营结果"],
        ["期内净现金流", metrics["net_cash_flow"], "含押金退还和资金往来，因此不等同利润"],
        ["押金/定金收款", metrics["deposit_receipts"], "客户资金性质，应持续跟踪后续退还或转收入"],
        ["退押/退定支出", metrics["refund_outflow"], "客户资金流出，不作为经营成本"],
    ]
    ws.append(["股东版摘要", "金额/比例", "解释"])
    for row in rows:
        ws.append(row)
    for r in range(2, 9):
        ws.cell(r, 2).number_format = '0.0%' if "率" in str(ws.cell(r, 1).value) else '#,##0.00;[Red]-#,##0.00;-'
    ws["A11"] = "股东沟通口径"
    ws["A11"].font = Font(bold=True, size=13)
    shareholder_notes = [
        "本月业务经营结果建议看营业收入、毛利和经营利润，不直接看收款总额。",
        "现金流受押金/定金收退影响明显，需和利润分开解释。",
        "当前报表基于财务部月度收支表，未做银行流水和逐单合同交叉核验，适合管理汇报，不作为审计报表。",
    ]
    for i, text in enumerate(shareholder_notes, 12):
        ws.cell(i, 1, text)
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=3)
    set_widths(ws, {"A": 26, "B": 18, "C": 90})
    style_sheet(ws)

    # Gross profit
    ws = wb.create_sheet("5月毛利表")
    gross_rows = [
        ["一、营业收入", metrics["revenue"], ""],
        ["租金", metrics["receipt_by_type"].get("租金", 0), "主表收款类型=租金"],
        ["经营性其他收款", metrics["operating_other_receipts"], "剔除老板转入后的其他收款"],
        ["二、直接成本", metrics["direct_cost"], ""],
        ["同行结算", metrics["payment_by_type"].get("同行结算", 0), ""],
        ["挂靠结算", metrics["payment_by_type"].get("挂靠结算", 0), ""],
        ["维修结算", metrics["payment_by_type"].get("维修结算", 0), ""],
        ["订单直接费用", metrics["direct_extra"], "从其他/报销中按摘要关键词归类"],
        ["三、毛利", metrics["gross_profit"], "营业收入-直接成本"],
        ["毛利率", metrics["gross_margin"], "毛利/营业收入"],
        ["四、期间费用", metrics["opex"], "工资、营销、办公、仲裁、电费等"],
        ["五、经营利润（现金口径）", metrics["operating_profit"], "毛利-期间费用"],
        ["经营利润率", metrics["operating_margin"], "经营利润/营业收入"],
    ]
    write_table(ws, 1, ["项目", "金额/比例", "备注"], gross_rows, [2])
    for r in range(2, 15):
        if "率" in str(ws.cell(r, 1).value):
            ws.cell(r, 2).number_format = "0.0%"
    set_widths(ws, {"A": 28, "B": 18, "C": 62})
    style_sheet(ws)

    # Income and payment summaries
    ws = wb.create_sheet("收入结构")
    income_rows = sorted(metrics["receipt_by_type"].items(), key=lambda x: -x[1])
    write_table(ws, 1, ["收款类型", "金额"], income_rows, [2])
    chart = PieChart()
    labels = Reference(ws, min_col=1, min_row=2, max_row=1 + len(income_rows))
    data = Reference(ws, min_col=2, min_row=1, max_row=1 + len(income_rows))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    chart.title = "收款结构"
    ws.add_chart(chart, "D2")
    set_widths(ws, {"A": 20, "B": 18, "D": 18})
    style_sheet(ws)

    ws = wb.create_sheet("支出结构")
    pay_rows = sorted(metrics["payment_by_type"].items(), key=lambda x: -x[1])
    write_table(ws, 1, ["付款类型", "金额"], pay_rows, [2])
    chart = BarChart()
    data = Reference(ws, min_col=2, min_row=1, max_row=1 + len(pay_rows))
    cats = Reference(ws, min_col=1, min_row=2, max_row=1 + len(pay_rows))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.title = "支出结构"
    chart.y_axis.title = "金额"
    ws.add_chart(chart, "D2")
    set_widths(ws, {"A": 22, "B": 18, "D": 18})
    style_sheet(ws)

    # Raw detail tabs
    ws = wb.create_sheet("收款明细")
    ws.append(["日期", "客户编号", "收款方式", "收款类型", "金额", "摘要", "备注"])
    for r in receipts:
        ws.append([r["date"], r["customer"], r["method"], r["type"], r["amount"], r["summary"], r["remark"]])
    set_widths(ws, {"A": 14, "B": 18, "C": 14, "D": 14, "E": 14, "F": 54, "G": 34})
    style_sheet(ws)
    for cell in ws["E"][1:]:
        cell.number_format = '#,##0.00;[Red]-#,##0.00;-'

    ws = wb.create_sheet("支出明细")
    ws.append(["日期", "付款对象", "支付方式", "付款类型", "金额", "报销摘要", "备注", "经办"])
    for p in payments:
        ws.append([p["date"], p["payee"], p["method"], p["type"], p["amount"], p["summary"], p["remark"], p["handler"]])
    set_widths(ws, {"A": 14, "B": 28, "C": 14, "D": 18, "E": 14, "F": 60, "G": 30, "H": 14})
    style_sheet(ws)
    for cell in ws["E"][1:]:
        cell.number_format = '#,##0.00;[Red]-#,##0.00;-'

    # Checks
    ws = wb.create_sheet("核对与风险")
    check_rows = [
        ["资金账户收款合计", metrics["cash_receipts"], 2249612.55, metrics["cash_receipts"] - 2249612.55, "OK" if abs(metrics["cash_receipts"] - 2249612.55) < 0.01 else "需复核"],
        ["收款类型合计（不含违章押金）", metrics["receipt_type_total_excl_violation_deposit"], 2228612.55, metrics["receipt_type_total_excl_violation_deposit"] - 2228612.55, "OK" if abs(metrics["receipt_type_total_excl_violation_deposit"] - 2228612.55) < 0.01 else "需复核"],
        ["支出明细合计", metrics["cash_payments"], 1346754.70, metrics["cash_payments"] - 1346754.70, "OK" if abs(metrics["cash_payments"] - 1346754.70) < 0.01 else "需复核"],
        ["老板转入剔除", metrics["boss_transfer"], metrics["boss_transfer"], 0, "已从营业收入剔除"],
        ["押金/定金剔除", metrics["deposit_receipts"], metrics["deposit_receipts"], 0, "未计入营业收入"],
        ["未细分期间费用", metrics["unclassified_extra"], metrics["unclassified_extra"], 0, "已保守计入期间费用"],
    ]
    write_table(ws, 1, ["检查项", "实际", "期望/口径", "差异", "状态"], check_rows, [2, 3, 4])
    ws["A9"] = "主要风险说明"
    ws["A9"].font = Font(bold=True)
    risks = [
        "1. 部分租金摘要包含“租金+押金”，虽然收款类型列归为租金，但逐单准确毛利仍需合同/订单明细核验。",
        "2. 定金部分可能后续转收入，本报表按谨慎原则未计入营业收入。",
        "3. 本报表未和银行/POS流水逐笔核对，适合内部经营汇报，不等同审计口径财务报表。",
    ]
    for i, text in enumerate(risks, 10):
        ws.cell(i, 1, text)
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=5)
    set_widths(ws, {"A": 26, "B": 18, "C": 18, "D": 18, "E": 22})
    style_sheet(ws)
    for r in range(2, 8):
        ws.cell(r, 5).fill = green_fill if "OK" in str(ws.cell(r, 5).value) or "已" in str(ws.cell(r, 5).value) else warn_fill

    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, datetime):
                    cell.number_format = "yyyy-mm-dd"
        sheet.sheet_view.showGridLines = False

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_FILE)


def main():
    source = find_source_workbook()
    receipts, payments = load_data(source)
    metrics = classify(receipts, payments)
    build_workbook(source, receipts, payments, metrics)
    print(OUTPUT_FILE)
    print(f"revenue={metrics['revenue']:.2f}")
    print(f"gross_profit={metrics['gross_profit']:.2f}")
    print(f"gross_margin={metrics['gross_margin']:.4f}")
    print(f"operating_profit={metrics['operating_profit']:.2f}")
    print(f"cash_receipts={metrics['cash_receipts']:.2f}")
    print(f"cash_payments={metrics['cash_payments']:.2f}")


if __name__ == "__main__":
    main()
